# -*- coding: utf-8 -*-
"""Copy of Text Mining trial 1.ipynb
Automatically generated by Colaboratory.
Original file is located at
    https://colab.research.google.com/drive/1M_NpNxmd4TB_IpXLIc4K7X-z7hX0u0cX
"""

# Libraries for text preprocessing
import pandas
import re
import nltk
import openpyxl as xl
import json

#nltk.download('stopwords')
from nltk.corpus import stopwords
nltk.download('stopwords')
from nltk.stem.porter import PorterStemmer
from nltk.tokenize import RegexpTokenizer
#nltk.download('wordnet') 
nltk.download('wordnet')
from nltk.stem.wordnet import WordNetLemmatizer

keyword_list=[]

wb = xl.load_workbook('libraryFile.xlsx')
sheet = wb['Sheet1']
#requestedLibrary = "Numpy"
#requestedLibrary = requestedLibrary.lower()
libInfo = {}
listdict=[]

for row in range(2 , sheet.max_row + 1):
    #print(sheet.cell(row, 1).value)
    libInfo[sheet.cell(row, 1).value] = sheet.cell(row, 2).value


# sample text for performing tokenization
text = "i want help about numpy and scipy and numpy and keras"

##Creating a list of stop words and adding custom stopwords
stop_words = set(stopwords.words("english"))
##Creating a list of custom stopwords
new_words = ["using", "show", "result", "large", "also", "iv", "one", "two", "new", "previously", "shown"]
stop_words = stop_words.union(new_words)

corpus = []
#Remove punctuations
#text = re.sub('[^a-zA-Z]', ' ', dataset['abstract1'][i])
    
#Convert to lowercase
text = text.lower()
    
#remove tags
text=re.sub("&lt;/?.*?&gt;"," &lt;&gt; ",text)
   
#remove special characters and digits
text=re.sub("(\\d|\\W)+"," ",text)
    
#Convert to list from string
text = text.split()
    
#Stemming
ps=PorterStemmer()
#Lemmatisation
lem = WordNetLemmatizer()
text = [lem.lemmatize(word) for word in text if not word in stop_words] 
text = " ".join(text)
corpus.append(text)

print(corpus)
#Text Preparation (Convert the words in the corpus to Tokens or vectors - Tokenization/Vectorization)

#Creating a vector of word counts 
from sklearn.feature_extraction.text import CountVectorizer
import re
cv=CountVectorizer(min_df=0.8,stop_words=stop_words, max_features=10000, ngram_range=(1,1))
X=cv.fit_transform(corpus)



list(cv.vocabulary_.keys())[:10]

#Most frequently occuring words
def get_top_n_words(corpus, n=None):
    vec = CountVectorizer().fit(corpus)
    bag_of_words = vec.transform(corpus)
    sum_words = bag_of_words.sum(axis=0) 
    words_freq = [(word, sum_words[0, idx]) for word, idx in      
                   vec.vocabulary_.items()]
    words_freq =sorted(words_freq, key = lambda x: x[1], 
                       reverse=True)
    return words_freq[:n]
    #print(words_freq)
#Convert most freq words to dataframe for plotting bar plot
top_words = get_top_n_words(corpus, n=20)
top_df = pandas.DataFrame(top_words)
top_df.columns=["Word", "Freq"]
#Barplot of most freq words
import seaborn as sns
sns.set(rc={'figure.figsize':(13,8)})
g = sns.barplot(x="Word", y="Freq", data=top_df)
g.set_xticklabels(g.get_xticklabels(), rotation=30)


from sklearn.feature_extraction.text import TfidfTransformer
 
tfidf_transformer=TfidfTransformer(smooth_idf=True,use_idf=True)
tfidf_transformer.fit(X)
# get feature names
feature_names=cv.get_feature_names()
 
#fetch document for which keywords needs to be extracted
for text1 in corpus:
  doc=text1
 
  #generate tf-idf for the given document
  tf_idf_vector=tfidf_transformer.transform(cv.transform([doc]))

#Function for sorting tf_idf in descending order
from scipy.sparse import coo_matrix
def sort_coo(coo_matrix):
    tuples = zip(coo_matrix.col, coo_matrix.data)
    return sorted(tuples, key=lambda x: (x[1], x[0]), reverse=True)
 
def extract_topn_from_vector(feature_names, sorted_items, topn=10):
   """get the feature names and tf-idf score of top n items"""
    
   #use only topn items from vector
   sorted_items = sorted_items[:topn]
 
   score_vals = []
   feature_vals = []
    
   # word index and corresponding tf-idf score
   for idx, score in sorted_items:
        
       #keep track of feature name and its corresponding score
       score_vals.append(round(score, 3))
       feature_vals.append(feature_names[idx])
 
   #create a tuples of feature,score
   #results = zip(feature_vals,score_vals)
   results= {}
   for idx in range(len(feature_vals)):
       results[feature_vals[idx]]=score_vals[idx]
    
   return results
#sort the tf-idf vectors by descending order of scores
sorted_items=sort_coo(tf_idf_vector.tocoo())
#extract only the top n; n here is 10
keywords=extract_topn_from_vector(feature_names,sorted_items,5)
 
# now print the results
print("\nAbstract:")
print(doc)

print("\nKeywords:")
for k in keywords:
    keyword_list.append(str(k))
    print(k,keywords[k])


print(keyword_list)

for k in keyword_list:
    for r in (libInfo):
        if k in r: 
            listdict.append({r : libInfo[r]})

print(listdict)
